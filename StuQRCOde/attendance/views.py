from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.db.models import Count, Avg
from datetime import datetime, timedelta
import qrcode
from io import BytesIO
import base64

# PDF and Excel imports
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook

from django.conf import settings

from .models import QRCode, AttendanceRecord, AttendanceStatistics
from .forms import QRCodeForm
from courses.models import Course
from accounts.models import Student, Lecturer, User

# QR Code management views
@login_required
def qr_code_list(request):
    """List all QR codes"""
    # Students should not access QR code management
    if request.user.role == 'student':
        messages.error(request, 'Access denied. Students cannot access QR code management.')
        return redirect('dashboard:home')
    
    qr_codes = QRCode.objects.all().select_related('course', 'created_by')
    
    # Filter by role
    if request.user.role == 'lecturer':
        try:
            lecturer_profile = request.user.lecturer_profile
            qr_codes = qr_codes.filter(course__lecturer=lecturer_profile)
        except User.lecturer_profile.RelatedObjectDoesNotExist:
            messages.warning(request, 'You do not have a lecturer profile. Please complete your profile first.')
            qr_codes = QRCode.objects.none()
    
    paginator = Paginator(qr_codes, 10)
    page_number = request.GET.get('page')
    qr_codes = paginator.get_page(page_number)
    
    context = {
        'qr_codes': qr_codes,
        'current_time': timezone.now(),
    }
    return render(request, 'attendance/qr_code_list.html', context)


@login_required
def qr_code_create(request):
    """Create new QR code"""
    if request.user.role not in ['admin', 'lecturer']:
        messages.error(request, 'Access denied.')
        return redirect('attendance:qr_code_list')
    
    if request.method == 'POST':
        form = QRCodeForm(request.POST, user=request.user)
        if form.is_valid():
            qr_code = form.save(commit=False)
            qr_code.created_by = request.user.lecturer_profile
            qr_code.save()
            messages.success(request, 'QR Code generated successfully!')
            return redirect('attendance:qr_code_detail', pk=qr_code.pk)
    else:
        form = QRCodeForm(user=request.user)
    
    return render(request, 'attendance/qr_code_form_fixed.html', {'form': form})


@login_required
def qr_code_detail(request, pk):
    """QR code detail view"""
    qr_code = get_object_or_404(QRCode, pk=pk)

    # Check permissions
    if request.user.role == 'lecturer' and qr_code.course.lecturer != request.user.lecturer_profile:
        messages.error(request, 'Access denied.')
        return redirect('attendance:qr_code_list')

    # Generate QR code image URL instead of inline base64 for better mobile performance
    # The actual image generation is handled by qr_code_generate view

    # Calculate attendance statistics
    present_count = qr_code.attendance_records.filter(status='present').count()
    late_count = qr_code.attendance_records.filter(status='late').count()

    # Create scan URL
    base_url = settings.BASE_URL.rstrip('/')
    scan_url = f"{base_url}/attendance/scan/{qr_code.id}/"

    context = {
        'qr_code': qr_code,
        'scan_url': scan_url,
        'present_count': present_count,
        'late_count': late_count,
        'current_time': timezone.now(),
    }
    return render(request, 'attendance/qr_code_detail.html', context)


@login_required
def qr_code_delete(request, pk):
    """Delete QR code"""
    qr_code = get_object_or_404(QRCode, pk=pk)

    # Check permissions
    if request.user.role not in ['admin', 'lecturer'] or (request.user.role == 'lecturer' and qr_code.course.lecturer != request.user.lecturer_profile):
        messages.error(request, 'Access denied.')
        return redirect('attendance:qr_code_list')

    if request.method == 'POST':
        qr_code.delete()
        messages.success(request, 'QR Code deleted successfully!')
        return redirect('attendance:qr_code_list')

    return render(request, 'attendance/qr_code_confirm_delete.html', {'qr_code': qr_code})


@login_required
def qr_code_generate(request, pk):
    """Generate QR code image with optimized size for mobile"""
    # Students should not download QR codes
    if request.user.role == 'student':
        messages.error(request, 'Access denied. Students cannot download QR codes.')
        return redirect('dashboard:home')

    qr_code = get_object_or_404(QRCode, pk=pk)

    # Check permissions
    if request.user.role == 'lecturer' and qr_code.course.lecturer != request.user.lecturer_profile:
        messages.error(request, 'Access denied.')
        return redirect('attendance:qr_code_list')

    # Generate QR code image
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )

    scan_url = f"{settings.BASE_URL.rstrip('/')}/attendance/scan/{qr_code.id}/"
    qr.add_data(scan_url)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')

    response = HttpResponse(buffer.getvalue(), content_type='image/png')

    # Check if this is a download request
    if request.GET.get('download') == 'true':
        response['Content-Disposition'] = f'attachment; filename="qr_code_{qr_code.id}.png"'
    else:
        response['Content-Disposition'] = 'inline'

    return response


# Attendance tracking views
@login_required
def scan_qr_code(request, qr_id):
    """Scan QR code for attendance"""
    qr_code = get_object_or_404(QRCode, pk=qr_id)
    now = timezone.now()
    
    # Check if QR code is valid with specific error messages
    if not qr_code.is_active:
        messages.error(request, 'This QR code is inactive.')
        return redirect('dashboard:home')
    elif now < qr_code.valid_from:
        messages.error(request, f'This QR code is not yet valid. It becomes active at {qr_code.valid_from}.')
        return redirect('dashboard:home')
    elif now > qr_code.valid_until:
        messages.error(request, f'This QR code has expired. It was valid until {qr_code.valid_until}.')
        return redirect('dashboard:home')
    
    # Check if user is a student
    if request.user.role != 'student':
        messages.error(request, 'Only students can mark attendance.')
        return redirect('dashboard:home')
    
    try:
        student = request.user.student_profile
        if not student.date_of_birth:  # Proxy for completion
            messages.info(request, 'Please complete your student profile.')
            return redirect('accounts:complete_student_profile')
    except Student.DoesNotExist:
        messages.info(request, 'Please complete your student profile.')
        return redirect('accounts:complete_student_profile')
    
    # Check if already marked attendance
    if AttendanceRecord.objects.filter(
        student=student,
        course=qr_code.course,
        date=timezone.now().date()
    ).exists():
        messages.info(request, 'Attendance already marked for today.')
        return redirect('dashboard:home')
    
    # Mark attendance
    attendance_record = AttendanceRecord.objects.create(
        student=student,
        course=qr_code.course,
        qr_code=qr_code,
        date=timezone.now().date(),
        time_in=timezone.now(),
        status='present',
        marked_by='qr_scan',
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    # Update attendance statistics
    stats, created = AttendanceStatistics.objects.get_or_create(
        student=student,
        course=qr_code.course
    )
    stats.attended_classes += 1
    stats.total_classes += 1
    stats.calculate_percentage()
    stats.save()
    
    messages.success(request, 'Attendance marked successfully!')
    return redirect('dashboard:home')


@login_required
def attendance_record_list(request):
    """List attendance records"""
    records = AttendanceRecord.objects.all().select_related('student', 'course', 'qr_code')
    
    # Filter by role
    if request.user.role == 'lecturer':
        try:
            lecturer_profile = request.user.lecturer_profile
            records = records.filter(course__lecturer=lecturer_profile)
        except User.lecturer_profile.RelatedObjectDoesNotExist:
            messages.warning(request, 'You do not have a lecturer profile. Please contact administrator.')
            records = AttendanceRecord.objects.none()
    elif request.user.role == 'student':
        # Students can only see their own records
        try:
            student_profile = request.user.student_profile
            records = records.filter(student=student_profile)
        except Student.DoesNotExist:
            messages.info(request, 'Please complete your student profile.')
            return redirect('accounts:complete_student_profile')
    
    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    records = paginator.get_page(page_number)
    
    context = {
        'records': records,
    }
    return render(request, 'attendance/attendance_record_list.html', context)


@login_required
def attendance_record_detail(request, pk):
    """Attendance record detail view"""
    record = get_object_or_404(AttendanceRecord, pk=pk)
    
    # Check permissions
    if request.user.role == 'student':
        # Students can only view their own records
        if record.student != request.user.student_profile:
            messages.error(request, 'Access denied.')
            return redirect('attendance:attendance_record_list')
    elif request.user.role == 'lecturer' and record.course.lecturer != request.user.lecturer_profile:
        messages.error(request, 'Access denied.')
        return redirect('attendance:attendance_record_list')
    
    context = {
        'record': record,
    }
    return render(request, 'attendance/attendance_record_detail.html', context)


# Dashboard views
@login_required
def lecturer_dashboard(request):
    """Lecturer dashboard with attendance statistics"""
    if request.user.role != 'lecturer':
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    try:
        lecturer = request.user.lecturer_profile
    except Lecturer.DoesNotExist:
        messages.error(request, 'Lecturer profile not found.')
        return redirect('accounts:complete_lecturer_profile')
    
    # Get courses taught by this lecturer
    courses = Course.objects.filter(lecturer=lecturer).select_related('semester')
    
    # Get attendance statistics for each course
    course_stats = []
    for course in courses:
        total_students = Student.objects.filter(
            attendance_records__course=course
        ).distinct().count()
        
        total_classes = AttendanceRecord.objects.filter(course=course).values('date').distinct().count()
        attended_classes = AttendanceRecord.objects.filter(course=course).count()
        
        if total_classes > 0:
            avg_attendance = (attended_classes / (total_classes * total_students)) * 100 if total_students > 0 else 0
        else:
            avg_attendance = 0
        
        course_stats.append({
            'course': course,
            'total_students': total_students,
            'total_classes': total_classes,
            'avg_attendance': round(avg_attendance, 2),
        })
    
    context = {
        'lecturer': lecturer,
        'course_stats': course_stats,
    }
    return render(request, 'attendance/lecturer_dashboard.html', context)





# Report generation views
@login_required
def course_attendance_report(request, course_id):
    """Generate attendance report for a course"""
    course = get_object_or_404(Course, pk=course_id)
    
    # Students should not access course attendance reports
    if request.user.role == 'student':
        messages.error(request, 'Access denied. Students cannot access course attendance reports.')
        return redirect('dashboard:home')
    
    # Check permissions
    if request.user.role == 'lecturer' and course.lecturer != request.user.lecturer_profile:
        messages.error(request, 'Access denied.')
        return redirect('attendance:attendance_reports')
    
    # Get attendance data
    students = Student.objects.filter(
        attendance_records__course=course
    ).distinct().select_related('user')
    
    attendance_data = []
    for student in students:
        records = AttendanceRecord.objects.filter(
            student=student,
            course=course
        ).order_by('date')
        
        total_classes = records.count()
        present_classes = records.filter(status='present').count()
        absent_classes = total_classes - present_classes
        percentage = (present_classes / total_classes * 100) if total_classes > 0 else 0
        
        attendance_data.append({
            'student': student,
            'total_classes': total_classes,
            'present_classes': present_classes,
            'absent_classes': absent_classes,
            'percentage': round(percentage, 2),
            'records': records,
        })
    
    context = {
        'course': course,
        'attendance_data': attendance_data,
    }
    
    # Check if PDF format is requested
    if request.GET.get('format') == 'pdf':
        # Students should not download PDF reports
        if request.user.role == 'student':
            messages.error(request, 'Access denied. Students cannot download attendance reports.')
            return redirect('dashboard:home')

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="course_attendance_report_{course.code}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(f"Course Attendance Report - {course.name}", styles['Title']))
        elements.append(Spacer(1, 12))

        # Course info
        course_info = [
            ['Course Code:', course.code],
            ['Course Name:', course.name],
            ['Lecturer:', course.lecturer.user.get_full_name()],
            ['Semester:', course.semester.name],
        ]
        table = Table(course_info)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 12))

        # Attendance data
        data = [['#', 'Student ID', 'Student Name', 'Total Classes', 'Present', 'Absent', 'Percentage', 'Status']]
        for i, data_item in enumerate(attendance_data, 1):
            status = 'Good' if data_item['percentage'] >= 75 else 'Fair' if data_item['percentage'] >= 50 else 'Poor'
            data.append([
                str(i),
                data_item['student'].student_id,
                data_item['student'].user.get_full_name(),
                str(data_item['total_classes']),
                str(data_item['present_classes']),
                str(data_item['absent_classes']),
                f"{data_item['percentage']}%" if data_item['total_classes'] > 0 else '0%',
                status
            ])
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)

        doc.build(elements)
        return response

    elif request.GET.get('format') == 'excel':
        # Students should not download Excel reports
        if request.user.role == 'student':
            messages.error(request, 'Access denied. Students cannot download attendance reports.')
            return redirect('dashboard:home')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="course_attendance_report_{course.code}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Title
        ws['A1'] = f"Course Attendance Report - {course.name}"
        ws['A2'] = f"Course Code: {course.code}"
        ws['A3'] = f"Course Name: {course.name}"
        ws['A4'] = f"Lecturer: {course.lecturer.user.get_full_name()}"
        ws['A5'] = f"Semester: {course.semester.name}"

        # Headers
        headers = ['#', 'Student ID', 'Student Name', 'Total Classes', 'Present', 'Absent', 'Percentage', 'Status']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=7, column=col_num, value=header)

        # Data
        for row_num, data_item in enumerate(attendance_data, 8):
            status = 'Good' if data_item['percentage'] >= 75 else 'Fair' if data_item['percentage'] >= 50 else 'Poor'
            ws.cell(row=row_num, column=1, value=row_num-7)
            ws.cell(row=row_num, column=2, value=data_item['student'].student_id)
            ws.cell(row=row_num, column=3, value=data_item['student'].user.get_full_name())
            ws.cell(row=row_num, column=4, value=data_item['total_classes'])
            ws.cell(row=row_num, column=5, value=data_item['present_classes'])
            ws.cell(row=row_num, column=6, value=data_item['absent_classes'])
            ws.cell(row=row_num, column=7, value=data_item['percentage'])
            ws.cell(row=row_num, column=8, value=status)

        wb.save(response)
        return response
    
    return render(request, 'attendance/course_attendance_report.html', context)


@login_required
def student_attendance_report(request, student_id):
    """Generate attendance report for a student"""
    student = get_object_or_404(Student, pk=student_id)
    
    # Check permissions
    if request.user.role == 'student':
        # Students can only view their own report
        if student != request.user.student_profile:
            messages.error(request, 'Access denied.')
            return redirect('dashboard:home')
    elif request.user.role == 'lecturer':
        # Lecturer can view any student's report
        pass
    
    # Get attendance data
    records = AttendanceRecord.objects.filter(
        student=student
    ).select_related('course', 'qr_code').order_by('-date')
    
    # Group by course
    course_data = {}
    for record in records:
        if record.course not in course_data:
            course_data[record.course] = {
                'records': [],
                'total_classes': 0,
                'present_classes': 0,
            }
        course_data[record.course]['records'].append(record)
        course_data[record.course]['total_classes'] += 1
        if record.status == 'present':
            course_data[record.course]['present_classes'] += 1
    
    # Calculate percentages
    for course, data in course_data.items():
        data['percentage'] = round((data['present_classes'] / data['total_classes'] * 100), 2) if data['total_classes'] > 0 else 0
    
    context = {
        'student': student,
        'course_data': course_data,
    }
    return render(request, 'attendance/student_attendance_report.html', context)


@login_required
def attendance_reports(request):
    """Main attendance reports dashboard"""
    if request.user.role not in ['admin', 'lecturer']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    # Get courses based on user role
    if request.user.role == 'admin':
        courses = Course.objects.all()
    else:  # lecturer
        try:
            lecturer_profile = request.user.lecturer_profile
            courses = Course.objects.filter(lecturer=lecturer_profile)
        except User.lecturer_profile.RelatedObjectDoesNotExist:
            messages.warning(request, 'You do not have a lecturer profile. Please complete your profile first.')
            return redirect('accounts:complete_lecturer_profile')
    
    # Get overall statistics
    total_courses = courses.count()
    total_students = Student.objects.filter(
        attendance_records__course__in=courses
    ).distinct().count()
    
    total_classes = AttendanceRecord.objects.filter(
        course__in=courses
    ).values('date', 'course').distinct().count()
    
    total_attendance = AttendanceRecord.objects.filter(
        course__in=courses,
        status='present'
    ).count()
    
    # Recent activity
    recent_records = AttendanceRecord.objects.filter(
        course__in=courses
    ).select_related('student', 'course').order_by('-date', '-time_in')[:10]
    
    context = {
        'courses': courses,
        'total_courses': total_courses,
        'total_students': total_students,
        'total_classes': total_classes,
        'total_attendance': total_attendance,
        'recent_records': recent_records,
    }
    return render(request, 'attendance/attendance_reports.html', context)


@login_required
def attendance_analytics(request):
    """Advanced attendance analytics dashboard"""
    if request.user.role not in ['admin', 'lecturer']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    # Get courses based on user role
    if request.user.role == 'admin':
        courses = Course.objects.all()
    else:  # lecturer
        try:
            lecturer_profile = request.user.lecturer_profile
            courses = Course.objects.filter(lecturer=lecturer_profile)
        except User.lecturer_profile.RelatedObjectDoesNotExist:
            messages.warning(request, 'You do not have a lecturer profile. Please complete your profile first.')
            return redirect('accounts:complete_lecturer_profile')
    
    # Get analytics data
    course_analytics = []
    for course in courses:
        records = AttendanceRecord.objects.filter(course=course)
        
        total_classes = records.values('date').distinct().count()
        total_students = Student.objects.filter(
            attendance_records__course=course
        ).distinct().count()
        
        present_records = records.filter(status='present')
        total_present = present_records.count()
        
        if total_classes > 0 and total_students > 0:
            avg_attendance = (total_present / (total_classes * total_students)) * 100
        else:
            avg_attendance = 0
        
        course_analytics.append({
            'course': course,
            'total_classes': total_classes,
            'total_students': total_students,
            'avg_attendance': round(avg_attendance, 2),
            'total_present': total_present,
        })
    
    # Get monthly trends
    from datetime import datetime, timedelta
    
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    daily_attendance = []
    current_date = start_date
    while current_date <= end_date:
        daily_records = AttendanceRecord.objects.filter(
            course__in=courses,
            date=current_date
        )
        
        daily_attendance.append({
            'date': current_date,
            'total_records': daily_records.count(),
            'present_count': daily_records.filter(status='present').count(),
        })
        current_date += timedelta(days=1)
    
    context = {
        'courses': courses,
        'course_analytics': course_analytics,
        'daily_attendance': daily_attendance,
    }
    return render(request, 'attendance/attendance_analytics.html', context)
